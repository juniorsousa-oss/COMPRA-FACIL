import os, io, re, unicodedata, difflib
from datetime import datetime, timezone
import pandas as pd
import requests
import streamlit as st

st.set_page_config(page_title="Compra Fácil", page_icon="🛒", layout="wide", initial_sidebar_state="collapsed")
SUPABASE_URL = st.secrets.get("SUPABASE_URL", os.getenv("SUPABASE_URL", "https://cuixazpxkvniqldmmnth.supabase.co")).rstrip("/")
SUPABASE_KEY = st.secrets.get("SUPABASE_KEY", os.getenv("SUPABASE_KEY", ""))

st.markdown("""<style>
#MainMenu,footer,header{visibility:hidden}.block-container{max-width:1150px;padding:1rem .9rem 5rem}
.brand{display:flex;align-items:center;gap:12px;margin:4px 0 18px}.brand-icon{width:48px;height:48px;border-radius:14px;background:#111827;display:flex;align-items:center;justify-content:center;font-size:25px}.brand-title{font-size:2rem;font-weight:800;color:#111827}.subtitle{margin:-14px 0 18px 60px;color:#6b7280}
.card{border:1px solid #e5e7eb;border-radius:16px;padding:14px 16px;background:white;margin-bottom:10px}.label{font-size:.78rem;color:#6b7280}.value{font-size:1.35rem;font-weight:750;color:#111827}.product-name{font-size:1.05rem;font-weight:750;color:#111827}.muted{font-size:.8rem;color:#6b7280}.done{text-decoration:line-through;color:#6b7280}.pill{display:inline-block;padding:5px 9px;border-radius:999px;font-size:.72rem;font-weight:700;background:#ecfdf5;color:#047857}.empty{border:1px dashed #d1d5db;border-radius:16px;padding:32px 16px;text-align:center;color:#6b7280;background:#fafafa}.stButton>button,.stDownloadButton>button{border-radius:11px;min-height:42px}
@media(max-width:700px){.block-container{padding:.7rem .65rem 4rem}.brand-title{font-size:1.65rem}.brand-icon{width:44px;height:44px;font-size:22px}.subtitle{margin-left:56px;font-size:.78rem}.card{padding:12px}.value{font-size:1.1rem}}
</style>""", unsafe_allow_html=True)

if not SUPABASE_KEY:
    st.error("Supabase não configurado. Informe SUPABASE_URL e SUPABASE_KEY nos Secrets do Streamlit Cloud.")
    st.stop()

HEADERS={"apikey":SUPABASE_KEY,"Authorization":f"Bearer {SUPABASE_KEY}","Content-Type":"application/json"}

def db(table, method="GET", params=None, data=None):
    r=requests.request(method,f"{SUPABASE_URL}/rest/v1/{table}",headers=HEADERS,params=params,json=data,timeout=25)
    if not r.ok: raise RuntimeError(f"Supabase {r.status_code}: {r.text[:700]}")
    return r.json() if r.text else []

def now(): return datetime.now(timezone.utc).isoformat()
def num(v):
    try:
        if pd.isna(v): return 0.0
        if isinstance(v,str):
            s=v.strip().replace("R$","").replace(" ","")
            if "," in s and "." in s: s=s.replace(".","").replace(",",".")
            elif "," in s: s=s.replace(",",".")
            v=s
        return float(v or 0)
    except: return 0.0

def money(v): return f"R$ {num(v):,.2f}".replace(",","X").replace(".",",").replace("X",".")
def norm(v):
    s=unicodedata.normalize("NFKD",str(v or "")).encode("ascii","ignore").decode().lower().strip()
    return re.sub(r"[^a-z0-9]+","_",s).strip("_")

def clear():
    for f in (get_products,get_current,get_budget,get_history): f.clear()

@st.cache_data(ttl=15,show_spinner=False)
def get_products(): return db("produtos",params={"select":"*","order":"nome.asc"})
@st.cache_data(ttl=4,show_spinner=False)
def get_current(): return db("lista_atual",params={"select":"*","order":"categoria.asc,nome_produto.asc"})
@st.cache_data(ttl=10,show_spinner=False)
def get_budget():
    r=db("estado_app",params={"select":"orcamento","id":"eq.1"})
    return num(r[0].get("orcamento")) if r else 0
@st.cache_data(ttl=15,show_spinner=False)
def get_history(): return db("compras",params={"select":"*","order":"data_compra.desc"})

def add_item(name,cat,unit,qty,price):
    db("lista_atual","POST",data={"nome_produto":name.strip(),"categoria":cat,"unidade":unit or "un.","quantidade":num(qty),"preco_estimado":num(price),"preco_unitario":0,"confirmado":False,"atualizado_em":now()}); clear()
def edit_item(i,**data): data["atualizado_em"]=now(); db("lista_atual","PATCH",params={"id":f"eq.{i}"},data=data); clear()
def remove_item(i): db("lista_atual","DELETE",params={"id":f"eq.{i}"}); clear()
def save_budget(v):
    data={"orcamento":num(v),"atualizado_em":now()}; r=db("estado_app",params={"select":"id","id":"eq.1"})
    db("estado_app","PATCH" if r else "POST",params={"id":"eq.1"} if r else None,data=data if r else {"id":1,**data}); clear()

def product_stats(product_id, latest_price=None, latest_qty=None):
    items=db("itens_compra",params={"select":"compra_id,preco_unitario,quantidade,criado_em","produto_id":f"eq.{product_id}"})
    valid=[x for x in items if num(x.get("preco_unitario"))>0]
    prices=[num(x.get("preco_unitario")) for x in valid]
    if latest_price is not None and num(latest_price)>0: last=num(latest_price)
    elif valid: last=num(sorted(valid,key=lambda x:str(x.get("criado_em", "")))[-1].get("preco_unitario"))
    else: last=0
    purchases={str(x.get("compra_id")) for x in items if x.get("compra_id") is not None}
    if latest_qty is not None: qty=num(latest_qty)
    elif valid:
        latest_item=sorted(valid,key=lambda x:str(x.get("criado_em", "")))[-1]; qty=num(latest_item.get("quantidade"))
    else: qty=1
    db("produtos","PATCH",params={"id":f"eq.{product_id}"},data={"ultimo_preco":last,"preco_medio":sum(prices)/len(prices) if prices else 0,"menor_preco":min(prices) if prices else 0,"maior_preco":max(prices) if prices else 0,"ultima_quantidade":qty or 1,"quantidade_compras":len(purchases),"atualizado_em":now()})

def rebuild_product_stats(products):
    items=db("itens_compra",params={"select":"produto_id,compra_id,preco_unitario,quantidade,criado_em"})
    grouped={}
    for x in items:
        pid=x.get("produto_id")
        if pid is None: continue
        grouped.setdefault(str(pid),[]).append(x)
    updated=0
    for p in products:
        pid=str(p.get("id"))
        if pid not in grouped: continue
        rows=grouped[pid]; valid=[x for x in rows if num(x.get("preco_unitario"))>0]; prices=[num(x.get("preco_unitario")) for x in valid]; ordered=sorted(rows,key=lambda x:str(x.get("criado_em", ""))); latest=ordered[-1] if ordered else None
        db("produtos","PATCH",params={"id":f"eq.{p['id']}"},data={"ultimo_preco":num(latest.get("preco_unitario")) if latest else 0,"preco_medio":sum(prices)/len(prices) if prices else 0,"menor_preco":min(prices) if prices else 0,"maior_preco":max(prices) if prices else 0,"ultima_quantidade":num(latest.get("quantidade")) if latest else 1,"quantidade_compras":len({str(x.get("compra_id")) for x in rows if x.get("compra_id") is not None}),"atualizado_em":now()})
        updated+=1
    clear(); return updated

def find_product_by_name(products,name):
    key=norm(name)
    for p in products:
        if norm(p.get("nome"))==key: return p
    return None

def similarity_score(a,b):
    na,nb=norm(a),norm(b)
    if not na or not nb: return 0
    ratio=difflib.SequenceMatcher(None,na,nb).ratio(); ta=set(na.split("_")); tb=set(nb.split("_")); overlap=len(ta&tb)/max(1,len(ta|tb))
    return max(ratio,0.65*ratio+0.35*overlap)

def similar_products(name,products,limit=3):
    scored=sorted(((similarity_score(name,p.get("nome","")),p) for p in products if p.get("nome")),key=lambda x:x[0],reverse=True)
    return [(score,p) for score,p in scored[:limit] if score>=0.42]

def create_or_update_product(name,cat,unit,price,qty,product_id=None):
    if product_id:
        product_stats(product_id,latest_price=price,latest_qty=qty); return product_id
    existing=db("produtos",params={"select":"*","nome":f"ilike.{name.strip()}"})
    if existing:
        pid=existing[0]["id"]; product_stats(pid,latest_price=price,latest_qty=qty); return pid
    row=db("produtos","POST",data={"nome":name.strip(),"categoria":cat or "Mercearia","unidade":unit or "un.","ultimo_preco":num(price),"preco_medio":num(price) if num(price)>0 else 0,"menor_preco":num(price) if num(price)>0 else 0,"maior_preco":num(price) if num(price)>0 else 0,"ultima_quantidade":num(qty) or 1,"quantidade_compras":0,"atualizado_em":now()})[0]
    return row["id"]

def save_product(name,cat,unit,price=0,qty=1):
    existing=find_product_by_name(get_products(),name)
    if existing:
        db("produtos","PATCH",params={"id":f"eq.{existing['id']}"},data={"categoria":cat or existing.get("categoria","Mercearia"),"unidade":unit or existing.get("unidade","un."),"atualizado_em":now()})
        return existing["id"]
    return create_or_update_product(name,cat,unit,price,qty)

def finish(items,budget):
    est=sum(num(x.get("quantidade"))*num(x.get("preco_estimado")) for x in items); real=sum(num(x.get("quantidade"))*num(x.get("preco_unitario")) for x in items if x.get("confirmado"))
    c=db("compras","POST",data={"orcamento":num(budget),"valor_estimado":est,"valor_real":real,"saldo":num(budget)-real,"quantidade_itens":len(items),"data_compra":now()})[0]; rows=[]
    for x in items:
        q=num(x.get("quantidade")); e=num(x.get("preco_estimado")); p=num(x.get("preco_unitario")); pid=save_product(x["nome_produto"],x.get("categoria","Mercearia"),x.get("unidade","un."),p or e,q)
        rows.append({"compra_id":c["id"],"produto_id":pid,"nome_produto":x["nome_produto"],"quantidade":q,"unidade":x["unidade"],"preco_estimado":e,"preco_unitario":p,"valor_total":q*p,"ultimo_preco":e,"variacao_preco":p-e,"confirmado":bool(x.get("confirmado"))})
    if rows:
        db("itens_compra","POST",data=rows)
        for pid in {r["produto_id"] for r in rows}: product_stats(pid)
    db("lista_atual","DELETE",params={"id":"gt.0"}); clear()

def delete_history_purchase(purchase_id):
    db("itens_compra","DELETE",params={"compra_id":f"eq.{purchase_id}"})
    db("compras","DELETE",params={"id":f"eq.{purchase_id}"})
    clear()

@st.dialog("Adicionar produto à lista de compras")
def add_to_list_dialog(products):
    names=[p.get("nome","") for p in products if p.get("nome")]
    if not names: st.info("Cadastre produtos primeiro na aba Produtos."); return
    selected=st.selectbox("Produto",names); qty=st.number_input("Quantidade",min_value=.001,value=1.,step=1.,format="%.2f")
    p=next((x for x in products if x.get("nome")==selected),None)
    if st.button("Adicionar à lista",type="primary",use_container_width=True) and p:
        add_item(selected,p.get("categoria","Mercearia"),p.get("unidade","un."),qty,num(p.get("ultimo_preco"))); st.rerun()

@st.dialog("Adicionar novo produto")
def new_product_dialog(products):
    cats=sorted(set(["Mercearia","Hortifruti","Carnes","Bebidas","Laticínios e ovos","Padaria","Congelados","Limpeza","Higiene pessoal","Casa e utilidades","Pet","Infantil","Saúde e farmácia"]+[p.get("categoria","Mercearia") for p in products]))
    with st.form("new_product_form"):
        name=st.text_input("Nome",placeholder="Ex.: Arroz branco"); cat=st.selectbox("Categoria",cats); unit=st.text_input("Unidade",value="un.",placeholder="un., kg, L, pct...")
        if st.form_submit_button("Cadastrar produto",type="primary",use_container_width=True):
            if not name.strip(): st.error("Informe o nome do produto.")
            elif not unit.strip(): st.error("Informe a unidade de medida.")
            elif any(norm(p.get("nome"))==norm(name) for p in products): st.error("Esse produto já está cadastrado.")
            else: save_product(name,cat,unit.strip(),0,1); st.rerun()

def import_products_excel(uploaded,existing):
    try: df=pd.read_excel(uploaded,dtype=object)
    except Exception as e: raise RuntimeError(f"Não consegui ler o Excel: {e}")
    if df.empty: raise RuntimeError("A planilha está vazia.")
    original=list(df.columns); normalized={norm(c):c for c in original}
    def col(*names):
        for n in names:
            if norm(n) in normalized: return normalized[norm(n)]
        return None
    c_nome=col("NOME","Nome","PRODUTO","Produto","Descrição","Item","Material"); c_cat=col("CATEGORIA","Categoria"); c_unit=col("UNIDADE","Unidade","UN","Und","Medida")
    if not c_nome or not c_cat or not c_unit: raise RuntimeError("Para importar produtos, o Excel precisa obrigatoriamente das 3 colunas: NOME, CATEGORIA e UNIDADE.")
    keys={norm(p.get("nome")) for p in existing if p.get("nome")}; seen=set(); new=[]; skip=0
    for _,r in df.iterrows():
        name=str(r.get(c_nome,"") or "").strip(); cat=str(r.get(c_cat,"") or "").strip(); unit=str(r.get(c_unit,"") or "").strip(); k=norm(name)
        if not name or not cat or not unit or k in keys or k in seen: skip+=1; continue
        seen.add(k); new.append({"nome":name,"categoria":cat,"unidade":unit,"ultimo_preco":0,"preco_medio":0,"menor_preco":0,"maior_preco":0,"ultima_quantidade":1,"quantidade_compras":0,"atualizado_em":now()})
    return new,skip

def insert_imported(rows):
    if rows: db("produtos","POST",data=rows)
    clear()

def recurring_products(products):
    counts={}; quantities={}
    for purchase in get_history():
        seen=set(); items=db("itens_compra",params={"select":"nome_produto,quantidade","compra_id":f"eq.{purchase['id']}"})
        for item in items:
            key=norm(item.get("nome_produto"))
            if not key or key in seen: continue
            seen.add(key); counts[key]=counts.get(key,0)+1; quantities.setdefault(key,[]).append(num(item.get("quantidade")))
    byname={norm(p.get("nome")):p for p in products if p.get("nome")}; result=[]
    for key,count in counts.items():
        if count>=2 and key in byname:
            qs=quantities[key]; result.append((byname[key],count,max(.001,round(sum(qs)/len(qs),2))))
    return sorted(result,key=lambda x:(-x[1],x[0].get("nome","").lower()))

def import_standard_list(products,current):
    rec=recurring_products(products); keys={norm(x.get("nome_produto")) for x in current}; added=0
    for p,count,qty in rec:
        k=norm(p.get("nome"))
        if k in keys: continue
        add_item(p.get("nome"),p.get("categoria","Mercearia"),p.get("unidade","un."),qty,num(p.get("ultimo_preco"))); keys.add(k); added+=1
    return added,len(rec)

@st.dialog("Confirmar compra do item")
def confirm_dialog(item):
    st.markdown(f"### {item['nome_produto']}"); st.caption(f"Quantidade: {num(item.get('quantidade')):g} {item.get('unidade','un.')}")
    price=st.number_input("Preço unitário pago",min_value=0.,value=num(item.get("preco_unitario")) or num(item.get("preco_estimado")),step=.01,format="%.2f"); st.metric("Total do item",money(num(item.get("quantidade"))*price)); a,b=st.columns(2)
    if a.button("Cancelar",use_container_width=True): st.rerun()
    if b.button("Confirmar",type="primary",use_container_width=True): edit_item(item["id"],preco_unitario=price,confirmado=True); st.rerun()

def history_rows():
    rows=[]
    for purchase in get_history():
        date_text=str(purchase.get("data_compra", ""))[:19].replace("T"," ")
        for x in db("itens_compra",params={"select":"*","compra_id":f"eq.{purchase['id']}","order":"id.asc"}):
            q=num(x.get("quantidade")); p=num(x.get("preco_unitario")); last=num(x.get("ultimo_preco")); rows.append({"DATA COMPRA":date_text,"PRODUTO":x.get("nome_produto",""),"VALOR UNITÁRIO":p,"QNT":q,"ÚLTIMO VALOR":last,"VALOR TOTAL":q*p})
    return rows

def history_export_bytes(history):
    out=io.BytesIO(); summary=[]; details=[]
    for purchase in history:
        date_text=str(purchase.get("data_compra", ""))[:19].replace("T"," ")
        summary.append({"DATA COMPRA":date_text,"ORÇAMENTO":num(purchase.get("orcamento")),"VALOR ESTIMADO":num(purchase.get("valor_estimado")),"VALOR REAL":num(purchase.get("valor_real")),"SALDO":num(purchase.get("saldo")),"QTD. ITENS":int(num(purchase.get("quantidade_itens")))})
        items=db("itens_compra",params={"select":"*","compra_id":f"eq.{purchase['id']}","order":"id.asc"})
        for x in items:
            q=num(x.get("quantidade")); p=num(x.get("preco_unitario")); details.append({"DATA COMPRA":date_text,"PRODUTO":x.get("nome_produto",""),"UNIDADE":x.get("unidade","un."),"QNT":q,"VALOR UNITÁRIO":p,"VALOR TOTAL":q*p,"ÚLTIMO VALOR":num(x.get("ultimo_preco")),"VARIAÇÃO":num(x.get("variacao_preco"))})
    with pd.ExcelWriter(out,engine="openpyxl") as w:
        pd.DataFrame(summary,columns=["DATA COMPRA","ORÇAMENTO","VALOR ESTIMADO","VALOR REAL","SALDO","QTD. ITENS"]).to_excel(w,index=False,sheet_name="Compras")
        pd.DataFrame(details,columns=["DATA COMPRA","PRODUTO","UNIDADE","QNT","VALOR UNITÁRIO","VALOR TOTAL","ÚLTIMO VALOR","VARIAÇÃO"]).to_excel(w,index=False,sheet_name="Itens")
        from openpyxl.styles import Font,PatternFill,Alignment
        for ws in w.book.worksheets:
            ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
            for c in ws[1]: c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="2F6F5E"); c.alignment=Alignment(horizontal="center")
            for col in ws.columns:
                letter=col[0].column_letter; ws.column_dimensions[letter].width=min(max(max(len(str(c.value or "")) for c in col)+2,12),32)
            headers=[str(c.value) for c in ws[1]]
            for row in ws.iter_rows(min_row=2):
                for idx,h in enumerate(headers):
                    if h in {"ORÇAMENTO","VALOR ESTIMADO","VALOR REAL","SALDO","VALOR UNITÁRIO","VALOR TOTAL","ÚLTIMO VALOR","VARIAÇÃO"}: row[idx].number_format='R$ #,##0.00'
                    elif h=="QNT": row[idx].number_format='0.00'
    out.seek(0); return out.getvalue()

def parse_history_excel(uploaded):
    try: df=pd.read_excel(uploaded,dtype=object)
    except Exception as e: raise RuntimeError(f"Não consegui ler o Excel: {e}")
    if df.empty: raise RuntimeError("A planilha está vazia.")
    original=list(df.columns); normalized={norm(c):c for c in original}
    def col(*names):
        for n in names:
            if norm(n) in normalized: return normalized[norm(n)]
        return None
    c_prod=col("PRODUTO","Produto","Nome"); c_unit=col("VALOR UNITÁRIO","Valor Unitário","Preço Unitário","Preço"); c_qty=col("QNT","Quantidade","Qtd","Qtde"); c_last=col("ÚLTIMO VALOR","Ultimo Valor","Último Preço","Ultimo Preco")
    if not c_prod or not c_unit or not c_qty: raise RuntimeError("O Excel precisa conter as colunas PRODUTO, VALOR UNITÁRIO e QNT.")
    rows=[]
    for _,r in df.iterrows():
        name=str(r.get(c_prod,"") or "").strip(); q=num(r.get(c_qty)); p=num(r.get(c_unit)); last=num(r.get(c_last)) if c_last else 0
        if name and q>0: rows.append({"name":name,"qty":q,"price":p,"last":last})
    if not rows: raise RuntimeError("Nenhum item válido foi encontrado no Excel.")
    return rows

def stage_history_import(uploaded,products):
    rows=parse_history_excel(uploaded); exact=[]; unmatched=[]
    for x in rows:
        p=find_product_by_name(products,x["name"])
        if p: exact.append({**x,"product":p,"action":"vincular"})
        else: unmatched.append({**x,"suggestions":similar_products(x["name"],products,3),"action":"novo","selected":None})
    st.session_state["pending_history"]={"rows":rows,"exact":exact,"unmatched":unmatched}

def commit_staged_history(pending):
    resolved=[{**item,"product":item["product"]} for item in pending["exact"]]
    products=get_products()
    for item in pending["unmatched"]:
        if item.get("action")=="vincular" and item.get("selected"):
            p=next((p for p in products if str(p.get("id"))==str(item["selected"])),None)
            if not p: raise RuntimeError(f"Produto selecionado para {item['name']} não foi encontrado.")
            resolved.append({**item,"product":p})
        else:
            resolved.append({**item,"product":None,"new_category":item.get("new_category","Mercearia") or "Mercearia","new_unit":item.get("new_unit","un.") or "un."})
    real=sum(x["qty"]*x["price"] for x in resolved); estimated=sum(x["qty"]*(x["last"] if x["last"]>0 else x["price"]) for x in resolved)
    c=db("compras","POST",data={"orcamento":real,"valor_estimado":estimated,"valor_real":real,"saldo":0,"quantidade_itens":len(resolved),"data_compra":now()})[0]; item_rows=[]; created=[]
    for x in resolved:
        if x.get("product"):
            p=x["product"]; pid=p["id"]; unit=p.get("unidade","un."); previous=x["last"] if x["last"]>0 else num(p.get("ultimo_preco"))
        else:
            unit=x.get("new_unit","un.") or "un."; pid=create_or_update_product(x["name"],x.get("new_category","Mercearia"),unit,x["price"],x["qty"]); previous=x["last"] if x["last"]>0 else 0; created.append(x["name"])
        item_rows.append({"compra_id":c["id"],"produto_id":pid,"nome_produto":x["name"],"quantidade":x["qty"],"unidade":unit,"preco_estimado":previous,"preco_unitario":x["price"],"valor_total":x["qty"]*x["price"],"ultimo_preco":previous,"variacao_preco":x["price"]-previous,"confirmado":True})
    if item_rows: db("itens_compra","POST",data=item_rows)
    for pid in {x["produto_id"] for x in item_rows}: product_stats(pid)
    clear(); st.session_state.pop("pending_history",None)
    return len(resolved),real,len(created)

@st.dialog("Revisar produtos do histórico")
def history_mapping_dialog(pending):
    st.write("Alguns itens do Excel não foram encontrados no cadastro. O sistema sugere produtos semelhantes; você pode vincular ou criar um novo produto.")
    products=get_products(); cats=sorted(set(["Mercearia","Hortifruti","Carnes","Bebidas","Laticínios e ovos","Padaria","Congelados","Limpeza","Higiene pessoal","Casa e utilidades","Pet","Infantil","Saúde e farmácia"]+[p.get("categoria","Mercearia") for p in products]))
    for idx,item in enumerate(pending["unmatched"]):
        st.markdown(f"**{item['name']}** — {item['qty']:g} × {money(item['price'])}")
        sims=item.get("suggestions",[])
        if sims: st.caption("Sugestões: " + ", ".join(f"{p.get('nome')} ({score:.0%})" for score,p in sims))
        options=[("novo","Criar novo produto")] + [(str(p["id"]),p["nome"]) for p in products]; labels={k:v for k,v in options}; keys=[k for k,_ in options]; default="novo"
        if sims and sims[0][0]>=0.62: default=str(sims[0][1]["id"])
        choice=st.selectbox("Destino",keys,index=keys.index(default) if default in keys else 0,format_func=lambda k:labels[k],key=f"hist_dest_{idx}")
        item["action"]="novo" if choice=="novo" else "vincular"; item["selected"]=None if choice=="novo" else choice
        if choice=="novo":
            c1,c2=st.columns(2); item["new_category"]=c1.selectbox("Categoria",cats,index=cats.index(item.get("new_category","Mercearia")) if item.get("new_category","Mercearia") in cats else 0,key=f"hist_cat_{idx}"); item["new_unit"]=c2.text_input("Unidade",value=item.get("new_unit","un."),key=f"hist_unit_{idx}")
        st.divider()
    a,b=st.columns(2)
    if a.button("Cancelar",use_container_width=True): st.session_state.pop("pending_history",None); st.rerun()
    if b.button("Confirmar importação",type="primary",use_container_width=True):
        try:
            n,total,created=commit_staged_history(pending); st.success(f"{n} itens importados. Total: {money(total)}. {created} produto(s) novo(s) criado(s)."); st.rerun()
        except Exception as e: st.error(f"Erro ao concluir a importação: {e}")

@st.dialog("Excluir histórico")
def delete_history_dialog(purchase):
    st.warning("Essa ação excluirá a compra e todos os itens dela do histórico. Ela não poderá ser desfeita.")
    st.markdown(f"**Total:** {money(purchase.get('valor_real'))}"); a,b=st.columns(2)
    if a.button("Cancelar",use_container_width=True): st.rerun()
    if b.button("Excluir definitivamente",type="primary",use_container_width=True): delete_history_purchase(purchase["id"]); st.rerun()

try:
    products=get_products(); current=get_current(); budget=get_budget(); history=get_history()
except Exception as e:
    st.error(f"Não consegui acessar o banco de dados.\n\n{e}"); st.stop()

st.markdown('<div class="brand"><div class="brand-icon">🛒</div><div class="brand-title">Compra Fácil</div></div>',unsafe_allow_html=True)
st.markdown('<div class="subtitle">Lista de compras, preços e histórico em um só lugar.</div>',unsafe_allow_html=True)
estimated=sum(num(x.get("quantidade"))*num(x.get("preco_estimado")) for x in current); real=sum(num(x.get("quantidade"))*num(x.get("preco_unitario")) for x in current if x.get("confirmado")); done=sum(1 for x in current if x.get("confirmado"))
cols=st.columns(4)
for col,label,value in zip(cols,["Orçamento","Estimado","Real confirmado","Saldo"],[money(budget),money(estimated),money(real),money(budget-real)]):
    with col: st.markdown(f'<div class="card"><div class="label">{label}</div><div class="value">{value}</div></div>',unsafe_allow_html=True)
if budget>0: st.progress(min(real/budget,1),text=f"Consumo confirmado: {real/budget:.0%}")
st.caption(f"Lista atual: {done} de {len(current)} itens confirmados")
buy,prod,hist,ana,config=st.tabs(["🛒 Compra","📦 Produtos","📜 Histórico","📊 Análises","⚙️ Configurações"])

with buy:
    st.subheader("Lista de compras")
    a,b=st.columns(2)
    with a:
        if st.button("Adicionar produto à lista de compras",type="primary",use_container_width=True,key="open_add_list"): add_to_list_dialog(products)
    with b:
        if st.button("Importar lista padrão",use_container_width=True,key="import_standard"):
            try:
                added,total=import_standard_list(products,current)
                if added: st.success(f"{added} produtos recorrentes adicionados à lista.")
                elif total: st.info("Todos os produtos recorrentes já estão na lista.")
                else: st.info("Ainda não há produtos recorrentes no histórico.")
                st.rerun()
            except Exception as e: st.error(f"Erro ao importar lista padrão: {e}")
    if not current: st.markdown('<div class="empty"><strong>Sua lista está vazia.</strong><br>Adicione um produto cadastrado acima.</div>',unsafe_allow_html=True)
    for item in current:
        ok=bool(item.get("confirmado")); total=num(item.get("quantidade"))*num(item.get("preco_estimado")); st.markdown('<div class="card">',unsafe_allow_html=True); a,b=st.columns([4,1])
        with a:
            cls="product-name done" if ok else "product-name"; st.markdown(f'<div class="{cls}">{item["nome_produto"]}</div>',unsafe_allow_html=True); st.markdown(f'<div class="muted">{item.get("categoria","Mercearia")} · {num(item.get("quantidade")):g} {item.get("unidade","un.")}</div>',unsafe_allow_html=True)
        with b: st.markdown('<span class="pill">✓ Confirmado</span>' if ok else f'<div style="text-align:right;font-weight:700">{money(total)}</div>',unsafe_allow_html=True)
        c1,c2,c3=st.columns([1.2,1.3,1])
        with c1:
            q=st.number_input("Qtd",min_value=.001,value=num(item.get("quantidade")) or 1.,step=1.,key=f"q{item['id']}",disabled=ok)
            if q!=num(item.get("quantidade")) and not ok: edit_item(item["id"],quantidade=q); st.rerun()
        with c2: st.caption("Preço estimado"); st.write(money(item.get("preco_estimado")))
        with c3:
            if not ok:
                if st.button("Confirmar",type="primary",key=f"c{item['id']}",use_container_width=True): confirm_dialog(item)
            else: st.caption("Preço pago"); st.write(money(item.get("preco_unitario")))
        d1,d2=st.columns([3,1])
        with d1: st.caption(f"Total pago: {money(num(item.get('quantidade'))*num(item.get('preco_unitario')))}" if ok else "Aguardando confirmação do preço")
        with d2:
            if st.button("Excluir",key=f"d{item['id']}",use_container_width=True): remove_item(item["id"]); st.rerun()
        st.markdown('</div>',unsafe_allow_html=True)
    st.divider(); st.markdown("### Fechamento"); new_budget=st.number_input("Orçamento / dinheiro disponível",min_value=0.,value=budget,step=10.,format="%.2f")
    a,b=st.columns(2)
    if a.button("Salvar orçamento",use_container_width=True): save_budget(new_budget); st.rerun()
    if b.button("Finalizar e salvar compra",type="primary",disabled=not bool(current),use_container_width=True):
        try: finish(current,budget); st.rerun()
        except Exception as e: st.error(f"Erro ao finalizar: {e}")

with prod:
    st.subheader("Produtos")
    if st.button("Adicionar novo produto",type="primary",use_container_width=True,key="open_new_product"): new_product_dialog(products)
    st.info("Importação incremental: produtos já cadastrados são ignorados. Para importar, o Excel precisa conter NOME, CATEGORIA e UNIDADE.")
    if st.button("Atualizar informações dos produtos pelo histórico",use_container_width=True,key="rebuild_product_stats"):
        try:
            updated=rebuild_product_stats(products); st.success(f"Informações atualizadas para {updated} produtos com histórico."); st.rerun()
        except Exception as e: st.error(f"Erro ao atualizar informações: {e}")
    with st.expander("Importar lista Excel",expanded=False):
        uploaded=st.file_uploader("Escolher arquivo Excel",type=["xlsx","xlsm"],key="excel_products")
        if uploaded is not None:
            try:
                imported,skipped=import_products_excel(uploaded,products); st.success(f"{len(imported)} produtos novos encontrados; {skipped} linhas ignoradas.")
                if imported:
                    st.dataframe(pd.DataFrame(imported)[["nome","categoria","unidade"]],use_container_width=True,hide_index=True)
                    if st.button("Cadastrar apenas os novos produtos",type="primary",use_container_width=True,key="import_confirm"): insert_imported(imported); st.rerun()
            except Exception as e: st.error(f"Erro na importação: {e}")
    search=st.text_input("Pesquisar produto",placeholder="Arroz, leite, sabão..."); rows=[p for p in products if not search.strip() or search.lower() in p.get("nome","").lower()]
    if rows: st.dataframe(pd.DataFrame([{"Produto":p.get("nome"),"Categoria":p.get("categoria"),"Unidade":p.get("unidade"),"Último":money(p.get("ultimo_preco")),"Médio":money(p.get("preco_medio")),"Menor":money(p.get("menor_preco")),"Maior":money(p.get("maior_preco")),"Compras":p.get("quantidade_compras",0)} for p in rows]),use_container_width=True,hide_index=True)
    else: st.info("Nenhum produto encontrado.")

with hist:
    st.subheader("Histórico de compras")
    st.markdown("### Exportação completa")
    if history:
        st.caption("Exporta toda a base histórica registrada, com um resumo das compras e o detalhamento de todos os itens.")
        st.download_button("Exportar histórico completo para Excel",data=history_export_bytes(history),file_name="historico_completo_compras.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)
    else: st.info("Ainda não há histórico para exportar.")
    with st.expander("Importar histórico do Excel",expanded=False):
        st.caption("Informe somente o Excel. Não é necessário informar orçamento nem data: o orçamento será igual ao valor total e a data será registrada automaticamente como data da importação.")
        hist_file=st.file_uploader("Escolher Excel do histórico",type=["xlsx","xlsm"],key="excel_history")
        if hist_file is not None:
            try:
                preview=parse_history_excel(hist_file); st.dataframe(pd.DataFrame(preview).rename(columns={"name":"PRODUTO","qty":"QNT","price":"VALOR UNITÁRIO","last":"ÚLTIMO VALOR"}),use_container_width=True,hide_index=True)
                if st.button("Verificar produtos e continuar",type="primary",use_container_width=True,key="history_import_prepare"):
                    stage_history_import(hist_file,products); st.rerun()
            except Exception as e: st.error(f"Erro na importação: {e}")
    pending=st.session_state.get("pending_history")
    if pending: history_mapping_dialog(pending)
    st.divider()
    if not history: st.markdown('<div class="empty">Nenhuma compra finalizada ainda.</div>',unsafe_allow_html=True)
    for p in history:
        date_text=str(p.get("data_compra",""))[:16].replace("T"," ")
        with st.expander(f"{date_text} · {money(p.get('valor_real'))} · {p.get('quantidade_itens',0)} itens"):
            a,b,c,d=st.columns(4); a.metric("Orçamento",money(p.get("orcamento"))); b.metric("Estimado",money(p.get("valor_estimado"))); c.metric("Real",money(p.get("valor_real"))); d.metric("Saldo",money(p.get("saldo")))
            if st.button("Excluir este histórico",key=f"delhist{p['id']}",use_container_width=True): delete_history_dialog(p)
            items=db("itens_compra",params={"select":"*","compra_id":f"eq.{p['id']}","order":"id.asc"})
            if items: st.dataframe(pd.DataFrame([{"Produto":x.get("nome_produto"),"Qtd":num(x.get("quantidade")),"Un":x.get("unidade"),"Estimado":money(x.get("preco_estimado")),"Pago":money(x.get("preco_unitario")),"Total":money(x.get("valor_total")),"Confirmado":"Sim" if x.get("confirmado") else "Não"} for x in items]),use_container_width=True,hide_index=True)

with ana:
    st.subheader("Análises")
    if history:
        total=sum(num(x.get("valor_real")) for x in history); est=sum(num(x.get("valor_estimado")) for x in history); a,b,c=st.columns(3); a.metric("Compras",len(history)); b.metric("Gasto acumulado",money(total)); c.metric("Diferença",money(total-est)); st.bar_chart(pd.DataFrame({"Estimado":[num(x.get("valor_estimado")) for x in history],"Real":[num(x.get("valor_real")) for x in history]}))
    else: st.markdown('<div class="empty">Finalize uma compra para gerar análises.</div>',unsafe_allow_html=True)

with config:
    st.subheader("Configurações"); st.success("Banco de dados conectado"); st.write(f"**Supabase:** `{SUPABASE_URL}`"); st.write("**Banco:** PostgreSQL / Supabase"); st.write("**Lista:** compartilhada entre dispositivos")
    if st.button("Atualizar dados agora",use_container_width=True): clear(); st.rerun()
    st.caption("A chave do Supabase deve permanecer no Secrets do Streamlit Cloud, nunca no código do GitHub.")
