import os, io, re, unicodedata, difflib
from datetime import datetime, timezone
import pandas as pd
import requests
import streamlit as st

st.set_page_config(page_title="Compra Fácil", page_icon="🛒", layout="wide", initial_sidebar_state="collapsed")
SUPABASE_URL=st.secrets.get("SUPABASE_URL",os.getenv("SUPABASE_URL","https://cuixazpxkvniqldmmnth.supabase.co")).rstrip("/")
SUPABASE_KEY=st.secrets.get("SUPABASE_KEY",os.getenv("SUPABASE_KEY",""))

st.markdown('''<style>#MainMenu,footer,header{visibility:hidden}.block-container{max-width:1150px;padding:1rem .9rem 5rem}.brand{display:flex;align-items:center;gap:12px;margin:4px 0 18px}.brand-icon{width:48px;height:48px;border-radius:14px;background:#111827;display:flex;align-items:center;justify-content:center;font-size:25px}.brand-title{font-size:2rem;font-weight:800;color:#111827}.subtitle{margin:-14px 0 18px 60px;color:#6b7280}.card{border:1px solid #e5e7eb;border-radius:16px;padding:14px 16px;background:white;margin-bottom:10px}.label{font-size:.78rem;color:#6b7280}.value{font-size:1.35rem;font-weight:750;color:#111827}.product-name{font-size:1.05rem;font-weight:750;color:#111827}.muted{font-size:.8rem;color:#6b7280}.done{text-decoration:line-through;color:#6b7280}.pill{display:inline-block;padding:5px 9px;border-radius:999px;font-size:.72rem;font-weight:700;background:#ecfdf5;color:#047857}.empty{border:1px dashed #d1d5db;border-radius:16px;padding:32px 16px;text-align:center;color:#6b7280;background:#fafafa}.stButton>button,.stDownloadButton>button{border-radius:11px;min-height:42px}@media(max-width:700px){.block-container{padding:.7rem .65rem 4rem}.brand-title{font-size:1.65rem}.brand-icon{width:44px;height:44px;font-size:22px}.subtitle{margin-left:56px;font-size:.78rem}.card{padding:12px}.value{font-size:1.1rem}}</style>''',unsafe_allow_html=True)
if not SUPABASE_KEY: st.error("Supabase não configurado. Informe SUPABASE_URL e SUPABASE_KEY nos Secrets do Streamlit Cloud."); st.stop()
HEADERS={"apikey":SUPABASE_KEY,"Authorization":f"Bearer {SUPABASE_KEY}","Content-Type":"application/json","Prefer":"return=representation"}
def db(table,method="GET",params=None,data=None):
    r=requests.request(method,f"{SUPABASE_URL}/rest/v1/{table}",headers=HEADERS,params=params,json=data,timeout=25)
    if not r.ok: raise RuntimeError(f"Supabase {r.status_code}: {r.text[:700]}")
    return r.json() if r.text else []
def now(): return datetime.now(timezone.utc).isoformat()
def num(v):
    try:
        if pd.isna(v): return 0.0
        if isinstance(v,str):
            s=v.strip().replace("R$","").replace(" ","")
            if "," in s and "." in s: s=s.replace(".","").replace(",",".")
            elif "," in s: s=s.replace(",",".")
            v=s
        return float(v or 0)
    except: return 0.0
def money(v): return f"R$ {num(v):,.2f}".replace(",","X").replace(".",",").replace("X",".")
def norm(v):
    s=unicodedata.normalize("NFKD",str(v or "")).encode("ascii","ignore").decode().lower().strip(); return re.sub(r"[^a-z0-9]+","_",s).strip("_")
def clear():
    for f in (get_products,get_current,get_budget,get_history): f.clear()
@st.cache_data(ttl=15,show_spinner=False)
def get_products(): return db("produtos",params={"select":"*","order":"nome.asc"})
@st.cache_data(ttl=4,show_spinner=False)
def get_current(): return db("lista_atual",params={"select":"*","order":"categoria.asc,nome_produto.asc"})
@st.cache_data(ttl=10,show_spinner=False)
def get_budget():
    r=db("estado_app",params={"select":"orcamento","id":"eq.1"}); return num(r[0].get("orcamento")) if r else 0
@st.cache_data(ttl=15,show_spinner=False)
def get_history(): return db("compras",params={"select":"*","order":"data_compra.desc"})
def add_item(name,cat,unit,qty,price):
    db("lista_atual","POST",data={"nome_produto":name.strip(),"categoria":cat,"unidade":unit or "un.","quantidade":num(qty),"preco_estimado":num(price),"preco_unitario":0,"confirmado":False,"atualizado_em":now()}); clear()
def edit_item(i,**data): data["atualizado_em"]=now(); db("lista_atual","PATCH",params={"id":f"eq.{i}"},data=data); clear()
def remove_item(i):
    db("lista_atual","DELETE",params={"id":f"eq.{i}"})
    remaining = db("lista_atual", params={"select":"id","id":"gt.0","limit":1})
    if not remaining:
        save_budget(0)
    else:
        clear()
def save_budget(v):
    data={"orcamento":num(v),"atualizado_em":now()}; r=db("estado_app",params={"select":"id","id":"eq.1"}); db("estado_app","PATCH" if r else "POST",params={"id":"eq.1"} if r else None,data=data if r else {"id":1,**data}); clear()
def product_stats(pid):
    items=db("itens_compra",params={"select":"compra_id,preco_unitario,quantidade,criado_em","produto_id":f"eq.{pid}"}); valid=[x for x in items if num(x.get("preco_unitario"))>0]; prices=[num(x.get("preco_unitario")) for x in valid]; ordered=sorted(items,key=lambda x:str(x.get("criado_em",""))); latest=ordered[-1] if ordered else None
    db("produtos","PATCH",params={"id":f"eq.{pid}"},data={"ultimo_preco":num(latest.get("preco_unitario")) if latest else 0,"preco_medio":sum(prices)/len(prices) if prices else 0,"menor_preco":min(prices) if prices else 0,"maior_preco":max(prices) if prices else 0,"ultima_quantidade":num(latest.get("quantidade")) if latest else 1,"quantidade_compras":len({str(x.get("compra_id")) for x in items}),"atualizado_em":now()})
def rebuild_product_stats(products):
    items=db("itens_compra",params={"select":"produto_id,compra_id,preco_unitario,quantidade,criado_em"}); grouped={}
    for x in items:
        if x.get("produto_id") is not None: grouped.setdefault(str(x["produto_id"]),[]).append(x)
    updated=0
    for p in products:
        rows=grouped.get(str(p.get("id"))); 
        if not rows: continue
        valid=[x for x in rows if num(x.get("preco_unitario"))>0]; prices=[num(x.get("preco_unitario")) for x in valid]; latest=sorted(rows,key=lambda x:str(x.get("criado_em","")))[-1]
        db("produtos","PATCH",params={"id":f"eq.{p['id']}"},data={"ultimo_preco":num(latest.get("preco_unitario")),"preco_medio":sum(prices)/len(prices) if prices else 0,"menor_preco":min(prices) if prices else 0,"maior_preco":max(prices) if prices else 0,"ultima_quantidade":num(latest.get("quantidade")) or 1,"quantidade_compras":len({str(x.get("compra_id")) for x in rows}),"atualizado_em":now()}); updated+=1
    clear(); return updated
def find_product(products,name):
    k=norm(name); return next((p for p in products if norm(p.get("nome"))==k),None)
def similarity(a,b):
    na,nb=norm(a),norm(b)
    if not na or not nb:return 0
    ratio=difflib.SequenceMatcher(None,na,nb).ratio(); ta=set(na.split("_")); tb=set(nb.split("_")); overlap=len(ta&tb)/max(1,len(ta|tb)); return max(ratio,.65*ratio+.35*overlap)
def suggestions(name,products,limit=3):
    scored=sorted(((similarity(name,p.get("nome","")),p) for p in products if p.get("nome")),key=lambda x:x[0],reverse=True); return [(s,p) for s,p in scored[:limit] if s>=.42]
def create_product(name,cat,unit,price,qty):
    exact=find_product(get_products(),name)
    if exact: pid=exact["id"]
    else: pid=db("produtos","POST",data={"nome":name.strip(),"categoria":cat or "Mercearia","unidade":unit or "un.","ultimo_preco":0,"preco_medio":0,"menor_preco":0,"maior_preco":0,"ultima_quantidade":num(qty) or 1,"quantidade_compras":0,"atualizado_em":now()})[0]["id"]
    return pid
def finish(items,budget):
    est=sum(num(x.get("quantidade"))*num(x.get("preco_estimado")) for x in items); real=sum(num(x.get("quantidade"))*num(x.get("preco_unitario")) for x in items if x.get("confirmado")); c=db("compras","POST",data={"orcamento":num(budget),"valor_estimado":est,"valor_real":real,"saldo":num(budget)-real,"quantidade_itens":len(items),"data_compra":now()})[0]; rows=[]
    for x in items:
        q=num(x.get("quantidade")); e=num(x.get("preco_estimado")); p=num(x.get("preco_unitario")); pid=create_product(x["nome_produto"],x.get("categoria","Mercearia"),x.get("unidade","un."),p or e,q); rows.append({"compra_id":c["id"],"produto_id":pid,"nome_produto":x["nome_produto"],"quantidade":q,"unidade":x["unidade"],"preco_estimado":e,"preco_unitario":p,"valor_total":q*p,"ultimo_preco":e,"variacao_preco":p-e,"confirmado":bool(x.get("confirmado"))})
    if rows: db("itens_compra","POST",data=rows); [product_stats(pid) for pid in {x["produto_id"] for x in rows}]
    db("lista_atual","DELETE",params={"id":"gt.0"}); clear()
def delete_history_purchase(pid): db("itens_compra","DELETE",params={"compra_id":f"eq.{pid}"}); db("compras","DELETE",params={"id":f"eq.{pid}"}); clear()

def parse_history(uploaded):
    try: df=pd.read_excel(uploaded,dtype=object)
    except Exception as e: raise RuntimeError(f"Não consegui ler o Excel: {e}")
    if df.empty: raise RuntimeError("A planilha está vazia.")
    n={norm(c):c for c in df.columns}; col=lambda *names: next((n[norm(x)] for x in names if norm(x) in n),None); cp=col("PRODUTO","Produto","Nome"); cu=col("VALOR UNITÁRIO","Valor Unitário","Preço Unitário","Preço"); cq=col("QNT","Quantidade","Qtd","Qtde"); cl=col("ÚLTIMO VALOR","Ultimo Valor","Último Preço","Ultimo Preco")
    if not cp or not cu or not cq: raise RuntimeError("O Excel precisa conter PRODUTO, VALOR UNITÁRIO e QNT.")
    rows=[]
    for _,r in df.iterrows():
        name=str(r.get(cp,"") or "").strip(); q=num(r.get(cq)); p=num(r.get(cu)); last=num(r.get(cl)) if cl else 0
        if name and q>0: rows.append({"name":name,"qty":q,"price":p,"last":last})
    if not rows: raise RuntimeError("Nenhum item válido foi encontrado no Excel.")
    return rows

def stage_history(uploaded,products):
    rows=parse_history(uploaded); exact=[]; unmatched=[]
    # Itens nao cadastrados repetidos sao consolidados antes das sugestoes.
    grouped={}
    for x in rows:
        p=find_product(products,x["name"])
        if p:
            exact.append({**x,"product":p})
            continue
        key=norm(x["name"])
        if key not in grouped:
            grouped[key]={**x,"duplicate_count":1}
        else:
            g=grouped[key]
            old_qty=num(g["qty"]); new_qty=num(x["qty"]); total_qty=old_qty+new_qty
            if total_qty:
                g["price"]=(old_qty*num(g["price"])+new_qty*num(x["price"])) / total_qty
            g["qty"]=total_qty
            if num(x.get("last"))>0: g["last"]=num(x.get("last"))
            g["duplicate_count"]=g.get("duplicate_count",1)+1
    for x in grouped.values():
        x.update({"suggestions":suggestions(x["name"],products),"action":"novo","selected":None,"new_category":"Mercearia","new_unit":"un."})
        unmatched.append(x)
    st.session_state["pending_history"]={"exact":exact,"unmatched":unmatched}

def commit_history(pending,action_all=None):
    products=get_products(); resolved=[]; ignored=0; created=0
    for x in pending["exact"]: resolved.append({**x,"product":x["product"]})
    for x in pending["unmatched"]:
        action=x.get("action","novo")
        if action_all and action_all!="individual":
            if action_all=="ignorar": action="ignorar"
            elif action_all=="novo": action="novo"
            elif action_all=="melhor":
                sims=x.get("suggestions",[]); action="vincular" if sims else "novo"; x["selected"]=str(sims[0][1]["id"]) if sims else None
        if action=="ignorar": ignored+=1; continue
        if action=="vincular":
            p=next((p for p in products if str(p.get("id"))==str(x.get("selected"))),None)
            if not p: raise RuntimeError(f"Produto selecionado para {x['name']} não foi encontrado.")
            resolved.append({**x,"product":p})
        else: resolved.append({**x,"product":None})
    if not resolved: st.session_state.pop("pending_history",None); return 0,0,0,ignored
    real=sum(x["qty"]*x["price"] for x in resolved); estimated=sum(x["qty"]*(x["last"] if x["last"]>0 else (num(x["product"].get("ultimo_preco")) if x.get("product") else x["price"])) for x in resolved)
    c=db("compras","POST",data={"orcamento":real,"valor_estimado":estimated,"valor_real":real,"saldo":0,"quantidade_itens":len(resolved),"data_compra":now()})[0]; rows=[]
    for x in resolved:
        if x.get("product"): p=x["product"]; pid=p["id"]; unit=p.get("unidade","un."); previous=x["last"] if x["last"]>0 else num(p.get("ultimo_preco"))
        else: unit=x.get("new_unit","un.") or "un."; pid=create_product(x["name"],x.get("new_category","Mercearia"),unit,x["price"],x["qty"]); previous=x["last"] if x["last"]>0 else 0; created+=1
        rows.append({"compra_id":c["id"],"produto_id":pid,"nome_produto":x["name"],"quantidade":x["qty"],"unidade":unit,"preco_estimado":previous,"preco_unitario":x["price"],"valor_total":x["qty"]*x["price"],"ultimo_preco":previous,"variacao_preco":x["price"]-previous,"confirmado":True})
    db("itens_compra","POST",data=rows)
    for pid in {x["produto_id"] for x in rows}: product_stats(pid)
    clear(); st.session_state.pop("pending_history",None); return len(resolved),real,created,ignored

@st.dialog("Revisar produtos do histórico")
def mapping_dialog(pending):
    products=get_products(); cats=sorted(set(["Mercearia","Hortifruti","Carnes","Bebidas","Laticínios e ovos","Padaria","Congelados","Limpeza","Higiene pessoal","Casa e utilidades","Pet","Infantil","Saúde e farmácia"]+[p.get("categoria","Mercearia") for p in products]))
    st.write("Foram encontrados itens que não estão cadastrados. Você pode decidir item a item ou aplicar uma decisão geral.")
    mode=st.radio("Modo de decisão",["Item a item","Aplicar decisão geral"],horizontal=True,key="hist_decision_mode")
    if mode=="Aplicar decisão geral":
        general=st.selectbox("Decisão para todos os itens divergentes",["Vincular todos à melhor sugestão","Cadastrar todos como novos produtos","Ignorar todos os itens divergentes"],key="hist_general_action")
        st.info("A opção de vincular todos usa automaticamente a melhor correspondência encontrada para cada item. Se um item não tiver sugestão suficiente, ele será cadastrado como novo.")
    else: general="individual"
    for i,x in enumerate(pending["unmatched"]):
        st.markdown(f"**{x['name']}** — {x['qty']:g} × {money(x['price'])}")
        sims=x.get("suggestions",[])
        if sims: st.caption("Sugestões: "+", ".join(f"{p.get('nome')} ({s:.0%})" for s,p in sims))
        if mode=="Item a item":
            actions=["Vincular a produto cadastrado","Cadastrar como novo produto","Ignorar este item"]
            default=0 if sims and sims[0][0]>=.62 else 1
            action=st.radio("Ação",actions,index=default,key=f"hist_action_{i}")
            if action==actions[0]:
                opts=[(str(p["id"]),p["nome"]) for p in products]; labels=dict(opts); keys=list(labels); default_id=str(sims[0][1]["id"]) if sims and sims[0][0]>=.62 else (keys[0] if keys else "")
                if keys:
                    selected=st.selectbox("Produto correspondente",keys,index=keys.index(default_id) if default_id in keys else 0,format_func=lambda k:labels[k],key=f"hist_dest_{i}"); x["action"]="vincular"; x["selected"]=selected
                else: x["action"]="novo"
            elif action==actions[1]:
                x["action"]="novo"; c1,c2=st.columns(2); x["new_category"]=c1.selectbox("Categoria",cats,index=cats.index(x.get("new_category","Mercearia")) if x.get("new_category","Mercearia") in cats else 0,key=f"hist_cat_{i}"); x["new_unit"]=c2.text_input("Unidade",value=x.get("new_unit","un."),key=f"hist_unit_{i}")
            else: x["action"]="ignorar"; x["selected"]=None
        else:
            x["action"]="novo" if general.startswith("Cadastrar") else ("ignorar" if general.startswith("Ignorar") else "vincular")
        st.divider()
    a,b=st.columns(2)
    if a.button("Cancelar",use_container_width=True): st.session_state.pop("pending_history",None); st.rerun()
    if b.button("Confirmar importação",type="primary",use_container_width=True):
        try:
            ga="individual" if mode=="Item a item" else ("melhor" if general.startswith("Vincular") else "novo" if general.startswith("Cadastrar") else "ignorar")
            n,total,created,ignored=commit_history(pending,ga); st.success(f"{n} itens importados. {created} produto(s) novo(s). {ignored} item(ns) ignorado(s). Total: {money(total)}"); st.rerun()
        except Exception as e: st.error(f"Erro ao concluir a importação: {e}")

@st.dialog("Adicionar produto à lista de compras")
def add_list_dialog(products):
    names=[p.get("nome","") for p in products if p.get("nome")];
    if not names: st.info("Cadastre produtos primeiro na aba Produtos."); return
    selected=st.selectbox("Produto",names); qty=st.number_input("Quantidade",min_value=.001,value=1.,step=1.,format="%.2f"); p=find_product(products,selected)
    if st.button("Adicionar à lista",type="primary",use_container_width=True) and p: add_item(selected,p.get("categoria","Mercearia"),p.get("unidade","un."),qty,num(p.get("ultimo_preco"))); st.rerun()
@st.dialog("Adicionar novo produto")
def new_product_dialog(products):
    cats=sorted(set(["Mercearia","Hortifruti","Carnes","Bebidas","Laticínios e ovos","Padaria","Congelados","Limpeza","Higiene pessoal","Casa e utilidades","Pet","Infantil","Saúde e farmácia"]+[p.get("categoria","Mercearia") for p in products]))
    with st.form("new_product_form"):
        name=st.text_input("Nome"); cat=st.selectbox("Categoria",cats); unit=st.text_input("Unidade",value="un.")
        if st.form_submit_button("Cadastrar produto",type="primary",use_container_width=True):
            if not name.strip(): st.error("Informe o nome.")
            elif not unit.strip(): st.error("Informe a unidade.")
            elif find_product(products,name): st.error("Esse produto já está cadastrado.")
            else: create_product(name,cat,unit,0,1); clear(); st.rerun()
def import_products_excel(uploaded,existing):
    try: df=pd.read_excel(uploaded,dtype=object)
    except Exception as e: raise RuntimeError(f"Não consegui ler o Excel: {e}")
    n={norm(c):c for c in df.columns}; col=lambda *names: next((n[norm(x)] for x in names if norm(x) in n),None); a,b,c=col("NOME","PRODUTO","Descrição"),col("CATEGORIA"),col("UNIDADE","UN","Und","Medida")
    if not a or not b or not c: raise RuntimeError("O Excel precisa obrigatoriamente de NOME, CATEGORIA e UNIDADE.")
    keys={norm(p.get("nome")) for p in existing}; seen=set(); rows=[]
    for _,r in df.iterrows():
        name=str(r.get(a,"") or "").strip(); cat=str(r.get(b,"") or "").strip(); unit=str(r.get(c,"") or "").strip(); k=norm(name)
        if name and cat and unit and k not in keys and k not in seen: seen.add(k); rows.append({"nome":name,"categoria":cat,"unidade":unit,"ultimo_preco":0,"preco_medio":0,"menor_preco":0,"maior_preco":0,"ultima_quantidade":1,"quantidade_compras":0,"atualizado_em":now()})
    return rows,len(df)-len(rows)
def history_export(history):
    out=io.BytesIO(); summary=[]; details=[]
    for p in history:
        d=str(p.get("data_compra",""))[:19].replace("T"," "); summary.append({"DATA COMPRA":d,"ORÇAMENTO":num(p.get("orcamento")),"VALOR ESTIMADO":num(p.get("valor_estimado")),"VALOR REAL":num(p.get("valor_real")),"SALDO":num(p.get("saldo")),"QTD. ITENS":int(num(p.get("quantidade_itens")))})
        for x in db("itens_compra",params={"select":"*","compra_id":f"eq.{p['id']}","order":"id.asc"}): details.append({"DATA COMPRA":d,"PRODUTO":x.get("nome_produto"),"UNIDADE":x.get("unidade","un."),"QNT":num(x.get("quantidade")),"VALOR UNITÁRIO":num(x.get("preco_unitario")),"VALOR TOTAL":num(x.get("valor_total")),"ÚLTIMO VALOR":num(x.get("ultimo_preco")),"VARIAÇÃO":num(x.get("variacao_preco"))})
    with pd.ExcelWriter(out,engine="openpyxl") as w:
        pd.DataFrame(summary,columns=["DATA COMPRA","ORÇAMENTO","VALOR ESTIMADO","VALOR REAL","SALDO","QTD. ITENS"]).to_excel(w,index=False,sheet_name="Compras"); pd.DataFrame(details,columns=["DATA COMPRA","PRODUTO","UNIDADE","QNT","VALOR UNITÁRIO","VALOR TOTAL","ÚLTIMO VALOR","VARIAÇÃO"]).to_excel(w,index=False,sheet_name="Itens")
        from openpyxl.styles import Font,PatternFill,Alignment
        for ws in w.book.worksheets:
            ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
            for c in ws[1]: c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="2F6F5E"); c.alignment=Alignment(horizontal="center")
    out.seek(0); return out.getvalue()
@st.dialog("Confirmar compra do item")
def confirm_dialog(item):
    st.markdown(f"### {item['nome_produto']}")
    st.caption(f"Quantidade: {num(item.get('quantidade')):g} {item.get('unidade','un.')}")
    price=st.number_input("Preço unitário pago",min_value=0.,value=num(item.get("preco_unitario")) or num(item.get("preco_estimado")),step=.01,format="%.2f")
    st.metric("Total do item",money(num(item.get("quantidade"))*price))
    a,b=st.columns(2)
    if a.button("Cancelar",use_container_width=True): st.session_state.pop("confirm_item",None); st.rerun()
    if b.button("Confirmar",type="primary",use_container_width=True): edit_item(item["id"],preco_unitario=price,confirmado=True); st.session_state.pop("confirm_item",None); st.rerun()

@st.dialog("Excluir histórico")
def delete_history_dialog(p):
    st.warning("Essa ação excluirá a compra e todos os itens dela do histórico. Ela não poderá ser desfeita."); a,b=st.columns(2)
    if a.button("Cancelar",use_container_width=True): st.rerun()
    if b.button("Excluir definitivamente",type="primary",use_container_width=True): delete_history_purchase(p["id"]); st.rerun()

try: products=get_products(); current=get_current(); budget=get_budget(); history=get_history()
except Exception as e: st.error(f"Não consegui acessar o banco de dados.\n\n{e}"); st.stop()
st.markdown('<div class="brand"><div class="brand-icon">🛒</div><div class="brand-title">Compra Fácil</div></div>',unsafe_allow_html=True); st.markdown('<div class="subtitle">Lista de compras, preços e histórico em um só lugar.</div>',unsafe_allow_html=True)
est=sum(num(x.get("quantidade"))*num(x.get("preco_estimado")) for x in current); real=sum(num(x.get("quantidade"))*num(x.get("preco_unitario")) for x in current if x.get("confirmado")); done=sum(bool(x.get("confirmado")) for x in current)
cols=st.columns(4)
for col,label,value in zip(cols,["Orçamento","Estimado","Real confirmado","Saldo"],[money(budget),money(est),money(real),money(budget-real)]):
    with col: st.markdown(f'<div class="card"><div class="label">{label}</div><div class="value">{value}</div></div>',unsafe_allow_html=True)
if budget>0: st.progress(min(real/budget,1),text=f"Consumo confirmado: {real/budget:.0%}")
st.caption(f"Lista atual: {done} de {len(current)} itens confirmados")
buy,prod,hist,ana,config=st.tabs(["🛒 Compra","📦 Produtos","📜 Histórico","📊 Análises","⚙️ Configurações"])

# Ajuste rápido do orçamento da compra em andamento.
if budget > 0:
    with st.popover("Corrigir orçamento", use_container_width=False):
        st.caption("Corrija o valor informado no início da compra.")
        if "budget_edit_value" not in st.session_state:
            st.session_state["budget_edit_value"] = budget
        new_budget = st.number_input(
            "Novo orçamento (R$)",
            min_value=0.0,
            value=float(st.session_state["budget_edit_value"]),
            step=10.0,
            format="%.2f",
            key="budget_edit_value",
        )
        if st.button("Salvar novo orçamento", type="primary", use_container_width=True, key="save_budget_edit"):
            if new_budget <= 0:
                st.error("Informe um orçamento maior que zero.")
            else:
                save_budget(new_budget)
                st.rerun()
if st.session_state.get("confirm_item"):
    confirm_dialog(st.session_state["confirm_item"])

with buy:
    st.subheader("Lista de compras"); a,b=st.columns(2)
    with a:
        if st.button("Adicionar produto à lista de compras",type="primary",use_container_width=True,key="open_add_list"): add_list_dialog(products)
    with b:
        if st.button("Importar lista padrão",use_container_width=True,key="import_standard"):
            try:
                rec={}; qs={}
                for p in history:
                    seen=set()
                    for x in db("itens_compra",params={"select":"nome_produto,quantidade","compra_id":f"eq.{p['id']}"}):
                        k=norm(x.get("nome_produto"));
                        if k and k not in seen: seen.add(k); rec[k]=rec.get(k,0)+1; qs.setdefault(k,[]).append(num(x.get("quantidade")))
                by={norm(p.get("nome")):p for p in products}; added=0
                for k,cnt in rec.items():
                    if cnt>=2 and k in by and not any(norm(x.get("nome_produto"))==k for x in current): p=by[k]; add_item(p["nome"],p.get("categoria","Mercearia"),p.get("unidade","un."),round(sum(qs[k])/len(qs[k]),2),num(p.get("ultimo_preco"))); added+=1
                st.success(f"{added} produtos recorrentes adicionados."); st.rerun()
            except Exception as e: st.error(f"Erro: {e}")
    if not current: st.markdown('<div class="empty"><strong>Sua lista está vazia.</strong></div>',unsafe_allow_html=True)
    for item in current:
        ok=bool(item.get("confirmado")); total=num(item.get("quantidade"))*num(item.get("preco_estimado")); st.markdown('<div class="card">',unsafe_allow_html=True); a,b=st.columns([4,1])
        with a: st.markdown(f'<div class="product-name {"done" if ok else ""}">{item["nome_produto"]}</div><div class="muted">{item.get("categoria","Mercearia")} · {num(item.get("quantidade")):g} {item.get("unidade","un.")}</div>',unsafe_allow_html=True)
        with b: st.markdown('<span class="pill">✓ Confirmado</span>' if ok else f'<div style="text-align:right;font-weight:700">{money(total)}</div>',unsafe_allow_html=True)
        c1,c2,c3=st.columns([1.2,1.3,1])
        with c1:
            q=st.number_input("Qtd",min_value=.001,value=num(item.get("quantidade")) or 1.,step=1.,key=f"q{item['id']}",disabled=ok)
            if q!=num(item.get("quantidade")) and not ok: edit_item(item["id"],quantidade=q); st.rerun()
        with c2: st.caption("Preço estimado"); st.write(money(item.get("preco_estimado")))
        with c3:
            if not ok:
                if st.button("Confirmar",type="primary",key=f"c{item['id']}",use_container_width=True):
                    st.session_state["confirm_item"]=item; st.rerun()
            else: st.caption("Preço pago"); st.write(money(item.get("preco_unitario")))
        d1,d2=st.columns([3,1]);
        with d1: st.caption(f"Total pago: {money(num(item.get('quantidade'))*num(item.get('preco_unitario')))}" if ok else "Aguardando confirmação do preço")
        with d2:
            if st.button("Excluir",key=f"d{item['id']}",use_container_width=True): remove_item(item["id"]); st.rerun()
        st.markdown('</div>',unsafe_allow_html=True)
    st.divider(); st.markdown("### Fechamento"); nb=st.number_input("Orçamento / dinheiro disponível",min_value=0.,value=budget,step=10.,format="%.2f"); a,b=st.columns(2)
    if a.button("Salvar orçamento",use_container_width=True): save_budget(nb); st.rerun()
    if b.button("Finalizar e salvar compra",type="primary",disabled=not bool(current),use_container_width=True):
        try: finish(current,budget); st.rerun()
        except Exception as e: st.error(f"Erro ao finalizar: {e}")
with prod:
    st.subheader("Produtos")
    if st.button("Adicionar novo produto",type="primary",use_container_width=True,key="open_new_product"): new_product_dialog(products)
    if st.button("Atualizar informações dos produtos pelo histórico",use_container_width=True,key="rebuild_product_stats"):
        try: st.success(f"Informações atualizadas para {rebuild_product_stats(products)} produtos."); st.rerun()
        except Exception as e: st.error(f"Erro: {e}")
    with st.expander("Importar lista Excel",expanded=False):
        up=st.file_uploader("Escolher arquivo Excel",type=["xlsx","xlsm"],key="excel_products")
        if up is not None:
            try:
                imp,skip=import_products_excel(up,products); st.success(f"{len(imp)} produtos novos; {skip} linhas ignoradas.")
                if imp and st.button("Cadastrar apenas os novos produtos",type="primary",use_container_width=True,key="import_confirm"): db("produtos","POST",data=imp); clear(); st.rerun()
            except Exception as e: st.error(f"Erro na importação: {e}")
    search=st.text_input("Pesquisar produto",placeholder="Arroz, leite, sabão..."); rows=[p for p in products if not search.strip() or search.lower() in p.get("nome","").lower()]
    if rows: st.dataframe(pd.DataFrame([{"Produto":p.get("nome"),"Categoria":p.get("categoria"),"Unidade":p.get("unidade"),"Último":money(p.get("ultimo_preco")),"Médio":money(p.get("preco_medio")),"Menor":money(p.get("menor_preco")),"Maior":money(p.get("maior_preco")),"Compras":p.get("quantidade_compras",0)} for p in rows]),use_container_width=True,hide_index=True)
with hist:
    st.subheader("Histórico de compras")
    if history: st.download_button("Exportar histórico completo para Excel",data=history_export(history),file_name="historico_completo_compras.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)
    else: st.info("Ainda não há histórico para exportar.")
    with st.expander("Importar histórico do Excel",expanded=False):
        st.caption("Não é necessário informar orçamento nem data. O orçamento será igual ao total e a data será registrada automaticamente."); hf=st.file_uploader("Escolher Excel do histórico",type=["xlsx","xlsm"],key="excel_history")
        if hf is not None:
            try:
                preview=parse_history(hf); st.dataframe(pd.DataFrame(preview).rename(columns={"name":"PRODUTO","qty":"QNT","price":"VALOR UNITÁRIO","last":"ÚLTIMO VALOR"}),use_container_width=True,hide_index=True)
                if st.button("Verificar produtos e continuar",type="primary",use_container_width=True,key="history_import_prepare"): stage_history(hf,products); st.rerun()
            except Exception as e: st.error(f"Erro na importação: {e}")
    pending=st.session_state.get("pending_history")
    if pending: mapping_dialog(pending)
    st.divider()
    if not history: st.markdown('<div class="empty">Nenhuma compra finalizada ainda.</div>',unsafe_allow_html=True)
    for p in history:
        d=str(p.get("data_compra",""))[:16].replace("T"," ")
        with st.expander(f"{d} · {money(p.get('valor_real'))} · {p.get('quantidade_itens',0)} itens"):
            a,b,c,d2=st.columns(4); a.metric("Orçamento",money(p.get("orcamento"))); b.metric("Estimado",money(p.get("valor_estimado"))); c.metric("Real",money(p.get("valor_real"))); d2.metric("Saldo",money(p.get("saldo")))
            if st.button("Excluir este histórico",key=f"delhist{p['id']}",use_container_width=True): delete_history_dialog(p)
            items=db("itens_compra",params={"select":"*","compra_id":f"eq.{p['id']}","order":"id.asc"});
            if items: st.dataframe(pd.DataFrame([{"Produto":x.get("nome_produto"),"Qtd":num(x.get("quantidade")),"Un":x.get("unidade"),"Estimado":money(x.get("preco_estimado")),"Pago":money(x.get("preco_unitario")),"Total":money(x.get("valor_total"))} for x in items]),use_container_width=True,hide_index=True)
with ana:
    st.subheader("Análises")
    if history:
        total=sum(num(x.get("valor_real")) for x in history); esth=sum(num(x.get("valor_estimado")) for x in history); a,b,c=st.columns(3); a.metric("Compras",len(history)); b.metric("Gasto acumulado",money(total)); c.metric("Diferença",money(total-esth)); st.bar_chart(pd.DataFrame({"Estimado":[num(x.get("valor_estimado")) for x in history],"Real":[num(x.get("valor_real")) for x in history]}))
    else: st.markdown('<div class="empty">Finalize uma compra para gerar análises.</div>',unsafe_allow_html=True)
with config:
    st.subheader("Configurações"); st.success("Banco de dados conectado"); st.write(f"**Supabase:** `{SUPABASE_URL}`"); st.write("**Banco:** PostgreSQL / Supabase"); st.write("**Lista:** compartilhada entre dispositivos");
    if st.button("Atualizar dados agora",use_container_width=True): clear(); st.rerun()
